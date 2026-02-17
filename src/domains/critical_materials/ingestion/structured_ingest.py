"""Structured ingestion utilities for critical materials datasets."""

import csv
import json
import os
from typing import Any, Dict, Iterable, List


SUPPORTED_STRUCTURED_EXTENSIONS = {".csv", ".json", ".jsonl", ".xlsx"}


class MissingDependencyError(RuntimeError):
    """Raised when an optional parser dependency is not installed."""


def _safe_column_name(value: Any, idx: int) -> str:
    text = str(value).strip() if value is not None else ""
    return text if text else f"column_{idx}"


def _iter_csv(path: str) -> Iterable[Dict[str, Any]]:
    with open(path, "r", newline="") as file:
        for row in csv.DictReader(file):
            if isinstance(row, dict):
                yield row


def _iter_json(path: str) -> Iterable[Dict[str, Any]]:
    with open(path, "r") as file:
        payload = json.load(file)
    if isinstance(payload, list):
        for item in payload:
            if isinstance(item, dict):
                yield item
        return
    if isinstance(payload, dict):
        yield payload


def _iter_jsonl(path: str) -> Iterable[Dict[str, Any]]:
    with open(path, "r") as file:
        for line in file:
            text = line.strip()
            if not text:
                continue
            item = json.loads(text)
            if isinstance(item, dict):
                yield item


def _iter_xlsx(path: str, sheet_name: str = None) -> Iterable[Dict[str, Any]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise MissingDependencyError(
            "openpyxl is required for .xlsx ingestion. Install with: python3 -m pip install openpyxl"
        ) from exc

    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet_names = [sheet_name] if sheet_name else list(workbook.sheetnames)

    for sheet in sheet_names:
        worksheet = workbook[sheet]
        iter_rows = worksheet.iter_rows(values_only=True)
        headers_row = next(iter_rows, None)
        if headers_row is None:
            continue
        headers = [_safe_column_name(value, idx) for idx, value in enumerate(headers_row, start=1)]
        for row in iter_rows:
            if row is None or not any(cell is not None and str(cell).strip() for cell in row):
                continue
            record = {headers[idx]: row[idx] if idx < len(row) else None for idx in range(len(headers))}
            record["__sheet_name"] = sheet
            yield record


def _has_required_fields(record: Dict[str, Any], required_fields: set) -> bool:
    if not required_fields:
        return True
    for field in required_fields:
        value = record.get(field)
        if value is None:
            return False
        if isinstance(value, str) and not value.strip():
            return False
    return True


def _normalize_record(
    record: Dict[str, Any],
    field_aliases: Dict[str, str],
    default_values: Dict[str, Any],
) -> Dict[str, Any]:
    normalized = dict(record)
    for alias, canonical in field_aliases.items():
        if alias in normalized and canonical not in normalized:
            normalized[canonical] = normalized[alias]
    for key, value in default_values.items():
        if key not in normalized or normalized[key] in ("", None):
            normalized[key] = value
    return normalized


def _project_record(record: Dict[str, Any], keep_fields: List[str]) -> Dict[str, Any]:
    if not keep_fields:
        return record

    projected = {field: record.get(field) for field in keep_fields}
    for meta_field in ("__source_path", "__source_type", "__sheet_name"):
        if meta_field in record and meta_field not in projected:
            projected[meta_field] = record[meta_field]
    return projected


def ingest_structured_sources(config: Dict[str, Any]) -> Dict[str, Any]:
    """Load tabular structured inputs from CSV, JSON, JSONL, and XLSX files."""
    config = config or {}
    source_paths = config.get("source_paths", [])
    required_fields = set(config.get("required_fields", []))
    field_aliases = dict(config.get("field_aliases", {}))
    default_values = dict(config.get("default_values", {}))
    include_source_metadata = bool(config.get("include_source_metadata", True))
    keep_fields = [str(field) for field in list(config.get("keep_fields", [])) if str(field).strip()]
    xlsx_sheet_name = config.get("xlsx_sheet_name")

    ingested: List[Dict[str, Any]] = []
    missing_paths: List[str] = []
    unsupported_paths: List[str] = []
    failed_paths: List[Dict[str, str]] = []
    missing_dependencies: List[str] = []
    skipped_required_count = 0

    for path in source_paths:
        if not os.path.exists(path):
            missing_paths.append(path)
            continue

        ext = os.path.splitext(path)[1].lower()
        if ext not in SUPPORTED_STRUCTURED_EXTENSIONS:
            unsupported_paths.append(path)
            continue

        try:
            if ext == ".csv":
                records = _iter_csv(path)
            elif ext == ".json":
                records = _iter_json(path)
            elif ext == ".jsonl":
                records = _iter_jsonl(path)
            else:
                records = _iter_xlsx(path, sheet_name=xlsx_sheet_name)
        except MissingDependencyError as exc:
            message = str(exc)
            if message not in missing_dependencies:
                missing_dependencies.append(message)
            failed_paths.append({"path": path, "error": message})
            continue
        except Exception as exc:  # pragma: no cover - protection for malformed files
            failed_paths.append({"path": path, "error": str(exc)})
            continue

        try:
            for record in records:
                normalized = _normalize_record(record, field_aliases, default_values)
                if include_source_metadata:
                    normalized["__source_path"] = path
                    normalized["__source_type"] = ext
                if not _has_required_fields(normalized, required_fields):
                    skipped_required_count += 1
                    continue
                ingested.append(_project_record(normalized, keep_fields))
        except Exception as exc:  # pragma: no cover - protection for malformed files
            failed_paths.append({"path": path, "error": str(exc)})
            continue

    status = "ok" if not missing_paths and not unsupported_paths and not failed_paths else "partial"
    return {
        "status": status,
        "records": ingested,
        "record_count": len(ingested),
        "missing_paths": missing_paths,
        "unsupported_paths": unsupported_paths,
        "failed_paths": failed_paths,
        "missing_dependencies": missing_dependencies,
        "skipped_required_count": skipped_required_count,
    }
